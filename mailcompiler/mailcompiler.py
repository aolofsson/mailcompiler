#!/usr/bin/env python3
"""mailcompiler (mc): build and query a contacts database.

Single CLI with no subcommand: the operation is inferred from the -i/-o
formats (taken from the file extension, or forced with --iformat/--oformat):
  mc -i MBOX|PST|VCF -o DB.json [...]      import/merge into the contacts DB
  mc -i CSV --iformat outlook -o DB.json   import an Outlook/Google CSV
  mc -i DB.json -o OUT.{csv,vcf} [...]     export matching records
  mc -i DB.json -o OUT.json --dedup        merge same-name contacts
  mc -i MBOX|PST -o OUT.jsonl --llm        dump a per-email JSONL corpus

JSON is the native database format. CSV and vCard are interchange formats:
a CSV/vCard is an import source or an export target, never a database to
operate on (so dedup/export read JSON). Outlook's CSV column layout is
selected with --iformat/--oformat outlook.

mbox import streams the file (only header blocks are parsed, bodies skipped),
so a 21 GB mbox is handled in one pass with minimal memory. PST import uses
pypff (libpff-python).

Spec (confirmed with user):
  - Contacts = recipients (To/Cc) of mail I SENT, plus senders (From) of mail I
    RECEIVED. Self addresses are excluded.
  - Spam-labeled messages are skipped (unsolicited, not correspondence).
  - Automated/bulk senders (no-reply, mailer-daemon, bulk ESP domains, ...) are
    filtered out.
  - Identities merged by display name (multiple addresses -> one person).
  - Company derived from email domain, blank for free providers.
"""

import argparse
import csv
import hashlib
import json
import os
import re
import sys
from collections import defaultdict
from datetime import date, timezone
from email.parser import HeaderParser, Parser
from email.header import decode_header, make_header
from email.utils import getaddresses, parsedate_to_datetime

import phonenumbers

SEP = re.compile(r'^From \d+@xxx ')

# Phone extraction from signatures.
DEFAULT_REGION = "US"          # assumed country for code-less numbers
BODY_CAP = 64 * 1024           # max body bytes captured per message
SIG_TAIL_LINES = 12            # signature = last N lines when no "-- " marker
_REPLY_MARKER = re.compile(
    r'^(on .*wrote:|-+\s*original message\s*-+|from:\s|sent from my )',
    re.IGNORECASE)
_HTML_TAG = re.compile(r'<[^>]+>')

FREE_PROVIDERS = {
    "gmail.com", "googlemail.com", "yahoo.com", "yahoo.co.uk", "ymail.com",
    "outlook.com", "hotmail.com", "live.com", "msn.com", "icloud.com",
    "me.com", "mac.com", "aol.com", "proton.me", "protonmail.com", "gmx.com",
    "gmx.net", "mail.com", "zoho.com", "yandex.com", "fastmail.com",
    "comcast.net", "verizon.net", "sbcglobal.net", "att.net", "qq.com",
    "163.com", "126.com",
}

# Local-part patterns that indicate an automated / non-human sender. Matched
# against a normalized local-part (._ collapsed to -).
BOT_LOCAL = re.compile(
    r'^(no-?reply|do-?not-?reply|donotreply|mailer-daemon|mailer|postmaster|'
    r'bounce|bounces|notification|notifications|notify|noreply|alert|alerts|'
    r'newsletter|newsletters|mailing|mailings|nepliy|automated|'
    r'auto-?reply|autoreply|root|daemon|cron)$'
)
BOT_SUBSTR = re.compile(
    r'(no-?reply|noreply|do-?not-?reply|donotreply|mailer-daemon|'
    r'marketing|newsletter|mailings?|unsub|spamproc)')
# Local-part prefixes (before a "+" tag) that indicate automated mail,
# e.g. GitHub's reply+<hash>@reply.github.com or notifications+<id>@...
BOT_TAG_PREFIX = {
    "reply", "notify", "notification", "notifications", "bounce", "bounces",
    "comment", "comments", "mention", "mentions", "noreply", "no-reply",
}

# Bulk email service provider / list domains (suffix or prefix match).
BOT_DOMAINS = (
    "mailchimp.com", "mcsv.net", "mcdlv.net", "rsgsv.net", "list-manage.com",
    "sendgrid.net", "sendgrid.com", "mailgun.org", "mailgun.com", "sparkpostmail.com",
    "amazonses.com", "mandrillapp.com", "sendinblue.com", "sib.email",
    "substack.com", "bounce.com", "bounces.google.com", "mailer.com",
    "constantcontact.com", "ccsend.com", "hubspotemail.net", "hs-send.com",
    "salesforce.com", "exct.net", "rs6.net", "klaviyomail.com", "intercom-mail.com",
    "mixmax.com", "mail.notion.so", "reply.github.com", "github.com",
    "mktomail.com", "beehiiv.com", "en25.com", "hubspotstarter.net",
    "connectedcommunity.org", "mailmarketo.com", "marketo.com",
    "notifications.", "noreply.", "engagement.", "marketing.", "email.",
    "news.", "mailing.", "updates.", "em.",
)


def dec(s):
    """Decode an RFC2047 header value to plain text."""
    if not s:
        return ""
    try:
        return str(make_header(decode_header(s)))
    except Exception:
        return s


def is_bot(email_addr):
    local, _, domain = email_addr.partition("@")
    if not domain:
        return True
    norm = local.lower().replace("_", "-").replace(".", "-")
    if BOT_LOCAL.match(norm) or BOT_SUBSTR.search(local.lower()):
        return True
    if "+" in local and local.split("+", 1)[0].lower() in BOT_TAG_PREFIX:
        return True
    for d in BOT_DOMAINS:
        if domain == d or domain.endswith("." + d) or domain.startswith(d):
            return True
    return False


def load_domain_list(path):
    """Read a domain list file (whitelist or blacklist) into a set of domains.

    One entry per line. Blank lines and lines starting with '#' are ignored, so
    the categorized list files with '# section' headers work unchanged. Entries
    may be written as "example.com" or "@example.com"; both mean the whole
    domain.
    """
    domains = set()
    with open(path, "r", encoding="utf-8") as fh:
        for line in fh:
            entry = line.strip()
            if not entry or entry.startswith("#"):
                continue
            domains.add(entry.lstrip("@").lower())
    return domains


def _domain_matches(domain, domain_set):
    """True if `domain` equals or is a subdomain of any entry in domain_set."""
    return any(domain == d or domain.endswith("." + d) for d in domain_set)


def is_blacklisted(email_addr, blacklist_domains):
    """True if the address's domain matches a blacklisted domain or subdomain."""
    if not blacklist_domains:
        return False
    return _domain_matches(email_addr.split("@")[-1].lower(), blacklist_domains)


def contact_domains(contact):
    """All email domains for a contact (primary_email plus every emails[] entry)."""
    domains = set()
    addrs = [contact.get("primary_email", "")] + list(contact.get("emails", []) or [])
    for addr in addrs:
        if addr and "@" in addr:
            domains.add(addr.split("@")[-1].lower())
    return domains


def contact_in_domains(contact, domain_set):
    """True if any of the contact's email domains matches the domain set."""
    if not domain_set:
        return False
    return any(_domain_matches(d, domain_set) for d in contact_domains(contact))


def clean(tok):
    """Strip stray wrapping quotes/punctuation from a name token."""
    return tok.strip(" '\"().,-")


def split_name(display, email_addr):
    """Return (first, last) from a display name, with email fallback."""
    name = (display or "").strip().strip('"\'').strip()
    # Cut org/department/handle suffixes leaked into the display name, e.g.
    # "Seo (Ethan)/dept/SUPEX" or "Name <tag>" -> keep the leading human part.
    name = re.split(r'\s*[(/\\<|]', name)[0].strip().rstrip(").-").strip()
    # No usable display name: synthesize from the email local-part, but only
    # from human-looking tokens (alpha, optional trailing digits, <=20 chars).
    # This recovers "eric.wallace.4" -> Eric Wallace while rejecting hash/unsub
    # local-parts like "8a795fa6-038a-4166-...".
    if not name or ("@" in name and " " not in name):
        local = email_addr.split("@")[0]
        toks = [t for t in re.split(r'[._\-+]+', local)
                if re.fullmatch(r'[A-Za-z]{2,}\d*', t) and len(t) <= 20]
        if len(toks) >= 2:
            return toks[0].capitalize(), toks[-1].capitalize()
        if len(toks) == 1:
            return toks[0].capitalize(), ""
        return "", ""
    if "," in name:  # "Last, First"
        last, _, first = name.partition(",")
        ftok = clean(first).split()
        return (ftok[0].capitalize() if ftok else ""), clean(last).capitalize()
    parts = clean(name).split()
    if not parts:                 # name was only punctuation (e.g. ".", "()")
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], parts[-1]


def company_from(email_addr):
    domain = email_addr.split("@")[-1].lower()
    if domain in FREE_PROVIDERS:
        return ""
    base = domain.split(".")
    if len(base) >= 2:
        return base[-2].capitalize()
    return domain.capitalize()


# ---- per-address accumulator ------------------------------------------------
class Rec:
    __slots__ = ("emails", "names", "phones", "num_sent", "num_recv",
                 "first", "last")

    def __init__(self):
        self.emails = defaultdict(int)   # email -> count
        self.names = defaultdict(int)    # display name -> count
        self.phones = defaultdict(int)   # E.164 phone -> count (from signatures)
        self.num_sent = 0                # I -> them
        self.num_recv = 0                # them -> I
        self.first = None
        self.last = None

    def touch(self, dt):
        if dt is None:
            return
        if self.first is None or dt < self.first:
            self.first = dt
        if self.last is None or dt > self.last:
            self.last = dt


# ---- CSV output / additive merge -------------------------------------------
# Legal values for the manual `type` column (blank = unset).
TYPE_VALUES = ["customer", "competitor", "investor", "reporter", "partner",
               "vendor", "other"]

CSV_FIELDS = ["type", "friend", "last_name", "first_name", "title", "company",
              "phone", "address", "primary_email", "emails", "num_emails",
              "num_sent", "num_received", "first_interaction",
              "last_interaction", "source", "linkedin", "import_date"]

# Source values may contain spaces (mbox filenames), so they are joined with
# this separator rather than a space (which the emails column uses).
SOURCE_SEP = " | "


def _to_int(v):
    try:
        return int(str(v).strip())
    except (TypeError, ValueError):
        return 0


def person_to_row(p, source):
    """Convert an in-memory person dict to a CSV row dict, tagged with source."""
    return {
        "type": "",
        "friend": "",
        "last_name": p["last_name"],
        "first_name": p["first_name"],
        "title": "",
        "company": p["company"],
        "phone": p.get("phone", ""),
        "address": "",
        "primary_email": p["primary_email"],
        "emails": list(p["emails"]),
        "num_emails": p["num_emails"],
        "num_sent": p["num_sent"],
        "num_received": p["num_received"],
        "first_interaction": p["first_interaction"],
        "last_interaction": p["last_interaction"],
        "source": source,
        "linkedin": "",
        "import_date": "",
    }


def _normalize_row(d):
    """Coerce a loaded record (from JSON or CSV) into the canonical row dict.

    Accepts `emails` as either a list (JSON) or a space-separated string (CSV).
    """
    emails = d.get("emails")
    if not isinstance(emails, list):
        emails = (emails or "").split()
    return {
        "type": str(d.get("type") or "").strip(),
        "friend": str(d.get("friend") or "").strip(),
        "last_name": str(d.get("last_name") or "").strip(),
        "first_name": str(d.get("first_name") or "").strip(),
        "title": str(d.get("title") or "").strip(),
        "company": str(d.get("company") or "").strip(),
        "phone": str(d.get("phone") or "").strip(),
        "address": str(d.get("address") or "").strip(),
        "primary_email": str(d.get("primary_email") or "").strip(),
        "emails": list(emails),
        "num_emails": _to_int(d.get("num_emails")),
        "num_sent": _to_int(d.get("num_sent")),
        "num_received": _to_int(d.get("num_received")),
        "first_interaction": (str(d.get("first_interaction") or "").strip() or None),
        "last_interaction": (str(d.get("last_interaction") or "").strip() or None),
        "source": str(d.get("source") or "").strip(),
        "linkedin": str(d.get("linkedin") or "").strip(),
        "import_date": str(d.get("import_date") or "").strip(),
    }


def load_rows(path):
    """Load a contacts file into a list of canonical row dicts.

    Format follows the extension: .json (the native store) or .csv. A missing or
    empty file yields an empty list.
    """
    if not os.path.isfile(path) or os.path.getsize(path) == 0:
        return []
    low = path.lower()
    if low.endswith(".json"):
        with open(path, "r", encoding="utf-8") as fh:
            text = fh.read().strip()
        data = json.loads(text) if text else []
        records = data if isinstance(data, list) else []
    elif low.endswith(".xlsx"):
        records = _read_xlsx(path)
    else:
        with open(path, "r", encoding="utf-8", newline="") as fh:
            records = list(csv.DictReader(fh))
    # A record is keyed by its email; LinkedIn-only contacts have no email but
    # are identified by their profile URL, so keep those too.
    return [_normalize_row(d) for d in records
            if str(d.get("primary_email") or "").strip()
            or str(d.get("linkedin") or "").strip()]


def _merge_date(a, b, newest):
    vals = [d for d in (a, b) if d]
    if not vals:
        return None
    return max(vals) if newest else min(vals)


def _union_sources(a, b):
    """Union two SOURCE_SEP-joined source strings, de-duped, order-preserving."""
    out = []
    for s in a.split(SOURCE_SEP) + b.split(SOURCE_SEP):
        s = s.strip()
        if s and s not in out:
            out.append(s)
    return SOURCE_SEP.join(out)


def merge_row(existing, new, force=False):
    """Merge `new` into `existing`. By default hand-edited text fields (type,
    name, company, ...) are preserved (the new value only fills a blank); with
    `force=True` a non-empty new value overwrites the existing one. Counts are
    overwritten with the latest import; emails and sources union; the date range
    widens."""
    for f in ("type", "friend", "last_name", "first_name", "title", "company",
              "phone", "address"):
        existing[f] = (new[f] or existing[f]) if force else (existing[f] or new[f])
    for e in new["emails"]:
        if e not in existing["emails"]:
            existing["emails"].append(e)
    existing["num_emails"] = new["num_emails"]
    existing["num_sent"] = new["num_sent"]
    existing["num_received"] = new["num_received"]
    existing["first_interaction"] = _merge_date(
        existing["first_interaction"], new["first_interaction"], newest=False)
    existing["last_interaction"] = _merge_date(
        existing["last_interaction"], new["last_interaction"], newest=True)
    existing["source"] = _union_sources(existing["source"], new["source"])
    nl, el = new.get("linkedin", ""), existing.get("linkedin", "")
    existing["linkedin"] = (nl or el) if force else (el or nl)
    # import_date reflects the most recent import that touched this record.
    existing["import_date"] = new.get("import_date") or existing.get("import_date", "")


def _write_atomic(path, write_fn):
    """Run write_fn(file) against a temp file then os.replace, so a crash
    mid-write cannot corrupt an existing contacts file we just merged into."""
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8", newline="") as out:
        write_fn(out)
    os.replace(tmp, path)


def _native_cells(r):
    """Row cells for the native full-column layout, in CSV_FIELDS order."""
    return [r.get("type", ""), r.get("friend", ""), r["last_name"],
            r["first_name"], r.get("title", ""), r["company"],
            r.get("phone", ""), r.get("address", ""), r["primary_email"],
            " ".join(r["emails"]), r["num_emails"], r["num_sent"],
            r["num_received"], r["first_interaction"] or "",
            r["last_interaction"] or "", r.get("source", ""),
            r.get("linkedin", ""), r.get("import_date", "")]


def write_csv_rows(path, rows):
    def _w(out):
        w = csv.writer(out)
        w.writerow(CSV_FIELDS)
        for r in rows:
            w.writerow(_native_cells(r))
    _write_atomic(path, _w)


def write_json_rows(path, rows):
    ordered = [{f: r.get(f) for f in CSV_FIELDS} for r in rows]

    def _w(out):
        json.dump(ordered, out, indent=2, ensure_ascii=False)
        out.write("\n")
    _write_atomic(path, _w)


# ---- xlsx backend (openpyxl) ------------------------------------------------
def _write_xlsx(path, header, rows_cells):
    """Write a single-sheet .xlsx with a header row + one row per record."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(header))
    for cells in rows_cells:
        ws.append(list(cells))
    tmp = path + ".tmp"
    wb.save(tmp)
    os.replace(tmp, path)


def _read_xlsx(path):
    """Read the first sheet of an .xlsx into a list of {header: str} dicts."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        rows_iter = wb.active.iter_rows(values_only=True)
        try:
            header = [("" if h is None else str(h)).strip() for h in next(rows_iter)]
        except StopIteration:
            return []
        out = []
        for row in rows_iter:
            d = {h: ("" if v is None else str(v))
                 for h, v in zip(header, row) if h}
            out.append(d)
        return out
    finally:
        wb.close()


def write_xlsx_rows(path, rows):
    """Write contacts in the native full-column layout as .xlsx."""
    _write_xlsx(path, CSV_FIELDS, (_native_cells(r) for r in rows))


# ---- Outlook CSV layout -----------------------------------------------------
# Microsoft Outlook's CSV import/export header names (the subset we map).
OUTLOOK_FIELDS = ["First Name", "Last Name", "Job Title", "Company",
                  "E-mail Address", "E-mail 2 Address", "E-mail 3 Address",
                  "Business Phone", "Business Street"]


def _outlook_cells(r):
    """Row cells for the Outlook layout (first three emails only)."""
    emails = r.get("emails") or []
    primary = r.get("primary_email") or ""
    ordered = ([primary] if primary else [])
    ordered += [e for e in emails if e and e != primary]
    e1, e2, e3 = (ordered + ["", "", ""])[:3]
    return [r.get("first_name", ""), r.get("last_name", ""),
            r.get("title", ""), r.get("company", ""),
            e1, e2, e3, r.get("phone", ""), r.get("address", "")]


def _outlook_row_from_dict(raw, source):
    """Map one Outlook header->value dict to a contact row (or None if empty)."""
    r = {(k or "").strip().lower(): (v or "").strip() for k, v in raw.items()}
    emails = [r.get(h, "") for h in
              ("e-mail address", "e-mail 2 address", "e-mail 3 address")]
    emails = [e for e in emails if e]
    addr = ", ".join(p for p in (
        r.get("business street", ""), r.get("business city", ""),
        r.get("business state", ""), r.get("business postal code", ""),
        r.get("business country", "")) if p)
    phone = r.get("business phone", "") or r.get("mobile phone", "")
    row = _normalize_row({
        "first_name": r.get("first name", ""),
        "last_name": r.get("last name", ""),
        "title": r.get("job title", ""),
        "company": r.get("company", ""),
        "phone": phone,
        "address": addr,
        "primary_email": emails[0] if emails else "",
        "emails": emails,
        "source": source,
    })
    if row["primary_email"] or row["first_name"] or row["last_name"]:
        return row
    return None


def write_outlook_csv(path, rows):
    """Write contacts in Outlook's CSV column layout (importable by Outlook and
    Google Contacts). Only the first three emails are kept (Outlook's limit)."""
    def _w(out):
        w = csv.writer(out)
        w.writerow(OUTLOOK_FIELDS)
        for r in rows:
            w.writerow(_outlook_cells(r))
    _write_atomic(path, _w)


def write_outlook_xlsx(path, rows):
    """Write contacts in Outlook's column layout as .xlsx."""
    _write_xlsx(path, OUTLOOK_FIELDS, (_outlook_cells(r) for r in rows))


def parse_outlook_csv(path):
    """Parse an Outlook-format CSV into contact row dicts."""
    source = os.path.basename(path)
    rows = []
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        for raw in csv.DictReader(fh):
            row = _outlook_row_from_dict(raw, source)
            if row:
                rows.append(row)
    return rows


def parse_outlook_xlsx(path):
    """Parse an Outlook-layout .xlsx into contact row dicts."""
    source = os.path.basename(path)
    rows = []
    for raw in _read_xlsx(path):
        row = _outlook_row_from_dict(raw, source)
        if row:
            rows.append(row)
    return rows


def parse_linkedin_csv(path):
    """Parse a LinkedIn 'Connections' CSV export into entry dicts.

    The export has a few-line 'Notes:' preamble before the real header row
    (First Name,Last Name,URL,Email Address,Company,Position,Connected On), so we
    scan for the header. Most rows have no email; the profile URL is the stable
    identifier. Returns a list of
    {first, last, url, email, company, position} dicts (Connected On unused).
    """
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        rows = list(csv.reader(fh))
    hidx = next((i for i, r in enumerate(rows)
                 if "First Name" in r and "Email Address" in r), None)
    if hidx is None:
        sys.exit("error: %s does not look like a LinkedIn Connections export "
                 "(no 'First Name'/'Email Address' header)" % path)
    col = {name: i for i, name in enumerate(rows[hidx])}

    def g(r, name):
        i = col.get(name)
        return r[i].strip() if i is not None and i < len(r) else ""

    entries = []
    for r in rows[hidx + 1:]:
        if not any(c.strip() for c in r):
            continue
        entries.append({
            "first": g(r, "First Name"),
            "last": g(r, "Last Name"),
            "url": g(r, "URL"),
            "email": g(r, "Email Address").lower(),
            "company": g(r, "Company"),
            "position": g(r, "Position"),
        })
    return entries


# ---- query / list helpers ---------------------------------------------------
def _csv_set(value):
    """Parse a comma-separated CLI value into a lowercased set, or None."""
    if value is None:
        return None
    return {v.strip().lower() for v in value.split(",") if v.strip()}


def _domain(addr):
    return addr.split("@")[-1].lower() if addr else ""


def build_criteria(args):
    """Turn parsed list args into a flat dict of active predicates."""
    return {
        "type": _csv_set(args.type),
        "company": _csv_set(args.company),
        "first_name": _csv_set(args.first_name),
        "last_name": _csv_set(args.last_name),
        "email_domain": _csv_set(args.email_domain),
        "min_emails": args.min_emails,
        "max_emails": args.max_emails,
        "min_sent": args.min_sent,
        "max_sent": args.max_sent,
        "min_received": args.min_received,
        "max_received": args.max_received,
        "last_after": args.last_after,
        "last_before": args.last_before,
        "first_after": args.first_after,
        "first_before": args.first_before,
    }


def matches(contact, crit):
    """Return True if a contact satisfies every active criterion (AND)."""
    # Text set-membership (case-insensitive exact).
    for field in ("type", "company", "first_name", "last_name"):
        allowed = crit[field]
        if allowed is not None and str(contact.get(field, "")).lower() not in allowed:
            return False
    if crit["email_domain"] is not None:
        if _domain(contact.get("primary_email", "")) not in crit["email_domain"]:
            return False

    # Numeric inclusive ranges.
    for field, lo, hi in (
        ("num_emails", crit["min_emails"], crit["max_emails"]),
        ("num_sent", crit["min_sent"], crit["max_sent"]),
        ("num_received", crit["min_received"], crit["max_received"]),
    ):
        val = contact.get(field, 0) or 0
        if lo is not None and val < lo:
            return False
        if hi is not None and val > hi:
            return False

    # Date inclusive ranges (ISO YYYY-MM-DD strings; null date fails any bound).
    for field, after, before in (
        ("last_interaction", crit["last_after"], crit["last_before"]),
        ("first_interaction", crit["first_after"], crit["first_before"]),
    ):
        if after is None and before is None:
            continue
        val = contact.get(field)
        if not val:
            return False
        if after is not None and val < after:
            return False
        if before is not None and val > before:
            return False

    return True


def load_domain_files(paths, label):
    """Load and union one or more domain-list files into a single set.

    Exits with an error if any path is missing or the union is empty.
    """
    domains = set()
    for path in paths:
        if not os.path.isfile(path):
            sys.exit("error: %s not found: %s" % (label, path))
        domains |= load_domain_list(path)
    if not domains:
        sys.exit("error: %s file(s) contain no domains: %s"
                 % (label, ", ".join(paths)))
    return domains


def load_domain_filters(args):
    """Load --whitelist / --blacklist into domain sets (None when not given).

    Each flag takes a list of files whose domains are unioned together.
    """
    whitelist = blacklist = None
    if getattr(args, "whitelist", None):
        whitelist = load_domain_files(args.whitelist, "whitelist")
    if getattr(args, "blacklist", None):
        blacklist = load_domain_files(args.blacklist, "blacklist")
    return whitelist, blacklist


def select_by_domains(contacts, whitelist, blacklist):
    """Filter contacts by email domain: keep only whitelisted, drop blacklisted.

    Matching is on any of a contact's email domains (primary_email plus every
    emails[] entry), and matches subdomains too. Returns
    (kept, n_dropped_not_whitelisted, n_dropped_blacklisted).
    """
    if whitelist is None and not blacklist:
        return list(contacts), 0, 0
    kept = []
    n_wl = n_bl = 0
    for c in contacts:
        if whitelist is not None and not contact_in_domains(c, whitelist):
            n_wl += 1
            continue
        if blacklist and contact_in_domains(c, blacklist):
            n_bl += 1
            continue
        kept.append(c)
    return kept, n_wl, n_bl


# ---- body / phone extraction ------------------------------------------------
def _decode_part(part):
    """Decode one non-multipart MIME part to text using its declared charset."""
    try:
        payload = part.get_payload(decode=True)
        if payload is None:
            return ""
        charset = part.get_content_charset() or "utf-8"
        return payload.decode(charset, "replace")
    except Exception:
        return ""


def _cap(text, cap):
    """Truncate text to `cap` bytes/chars (cap=None means no limit)."""
    return text if cap is None else text[:cap]


def _message_text(m, cap=BODY_CAP):
    """Return a message's plain-text body (HTML stripped if needed), capped.

    Uses the legacy email API (walk over parts) so it is lenient with the messy,
    possibly truncated messages produced by the capped mbox capture. `cap=None`
    keeps the full text.
    """
    plain = html = ""
    try:
        for part in m.walk():
            if part.get_content_maintype() == "multipart":
                continue
            if part.get_content_disposition() == "attachment":
                continue
            ctype = part.get_content_type()
            if ctype == "text/plain" and not plain:
                plain = _decode_part(part)
            elif ctype == "text/html" and not html:
                html = _decode_part(part)
    except Exception:
        pass
    if plain:
        return _cap(plain, cap)
    if html:
        return _cap(_HTML_TAG.sub(" ", html), cap)
    return ""


def _fresh_body(body):
    """Return the body up to the first quoted/reply line (the new message text,
    including its signature), dropping the quoted thread history below it."""
    if not body:
        return ""
    out = []
    for ln in body.splitlines():
        if ln.strip().startswith(">") or _REPLY_MARKER.match(ln.strip()):
            break
        out.append(ln)
    return "\n".join(out).strip()


def _signature_text(body):
    """Return just the signature region: the fresh (non-quoted) tail of a body,
    after a '-- ' delimiter if present, else its last SIG_TAIL_LINES lines."""
    if not body:
        return ""
    lines = body.splitlines()
    cut = len(lines)
    for i, ln in enumerate(lines):
        s = ln.strip()
        if s.startswith(">") or _REPLY_MARKER.match(s):
            cut = i
            break
    top = lines[:cut]
    sig_start = 0
    for i, ln in enumerate(top):
        if ln.strip() == "--":
            sig_start = i + 1
    sig = top[sig_start:] if sig_start else top[-SIG_TAIL_LINES:]
    return "\n".join(sig)


def _extract_phones(body, region=DEFAULT_REGION):
    """Return de-duped E.164 phone numbers found in a body's signature region."""
    out = []
    for line in _signature_text(body).splitlines():
        if "fax" in line.lower():
            continue
        try:
            matches = phonenumbers.PhoneNumberMatcher(line, region)
            for match in matches:
                if phonenumbers.is_valid_number(match.number):
                    e164 = phonenumbers.format_number(
                        match.number, phonenumbers.PhoneNumberFormat.E164)
                    if e164 not in out:
                        out.append(e164)
        except Exception:
            continue
    return out


# ---- message ingestion ------------------------------------------------------
# Each source (mbox / PST) yields a normalized message dict:
#   {"from": [(name, addr), ...], "to": [(name, addr), ...]  (To+Cc),
#    "date": <naive-UTC datetime or None>, "is_sent": bool, "is_spam": bool,
#    "self_hints": [addr, ...]}  -- addresses known to be the account owner.

# Outlook folder names that mark sent / spam mail (case-insensitive).
PST_SENT_FOLDERS = {"sent items", "sent", "sent mail"}
PST_SPAM_FOLDERS = {"junk email", "junk e-mail", "junk", "spam"}

# MAPI property tags used by the PST fallback when there are no RFC822 headers.
_MAPI_SENDER_NAME = 0x0C1A
_MAPI_SENDER_SMTP = 0x5D01
_MAPI_SENDER_EMAIL = 0x0C1F


def _normalize_dt(dt):
    """Coerce a datetime to naive UTC (or return None)."""
    if dt is None:
        return None
    try:
        if dt.tzinfo is not None:
            dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
    except Exception:
        return None
    return dt


def _ingest_message(msg, recs, self_set, include_cc=True):
    """Fold one normalized message into `recs`/`self_set`.

    Returns False if the message was skipped (spam), True otherwise. By default
    the other To/Cc recipients of mail I received (people on a thread with me,
    not the sender) are also recorded so they become contacts, but they do not
    count as sent or received -- only direct correspondence increments those.
    Pass include_cc=False (the --no-cc flag) to skip those co-recipients.
    """
    if msg["is_spam"]:
        return False
    for a in msg["self_hints"]:
        if a:
            self_set.add(a.lower())

    from_emails = [addr.lower() for _, addr in msg["from"] if addr]
    sent_by_me = msg["is_sent"] or any(a in self_set for a in from_emails)
    dt = msg["date"]
    # Each group is (pairs, count-attr or None, signature-phones-to-credit).
    if sent_by_me:
        for a in from_emails:        # learn self addresses
            self_set.add(a)
        groups = [(msg["to"], "num_sent", [])]   # never trust our own signature
    else:
        phones = _extract_phones(msg.get("body", ""))
        groups = [(msg["from"], "num_recv", phones)]
        if include_cc:
            # Co-recipients on mail I received (To/Cc minus the sender): include
            # them as contacts but credit no count and no signature phone.
            seen = set(from_emails)
            cc_pairs = [(n, a) for (n, a) in msg["to"]
                        if (a or "").lower().strip() not in seen]
            groups.append((cc_pairs, None, []))

    for pairs, attr, phones in groups:
        for raw_name, addr in pairs:
            addr = (addr or "").lower().strip()
            if not addr or "@" not in addr or addr in self_set:
                continue
            r = recs[addr]
            r.emails[addr] += 1
            nm = dec(raw_name).strip()
            if nm and "@" not in nm:
                r.names[nm] += 1
            for ph in phones:        # signature phones (received mail only)
                r.phones[ph] += 1
            if attr:                 # co-recipients (attr None) add no count
                setattr(r, attr, getattr(r, attr) + 1)
            r.touch(dt)
    return True


def iter_mbox_messages(path, body_cap=BODY_CAP):
    """Yield normalized messages from a Gmail Takeout mbox (streamed).

    Each message block is captured up to `body_cap` bytes (headers + body) and
    fully MIME-parsed so a plain-text body is available. The default cap keeps
    only the start of the body (where the signature lives) to bound memory/CPU;
    `body_cap=None` captures the whole message (full body, e.g. for --llm).
    """
    parser = Parser()

    def normalize(block):
        m = parser.parsestr("".join(block))
        labels = {x.strip() for x in (m.get("X-Gmail-Labels") or "").split(",")}
        try:
            dt = _normalize_dt(parsedate_to_datetime(m.get("Date")))
        except Exception:
            dt = None
        return {
            "subject": dec(m.get("Subject") or ""),
            "from": getaddresses(m.get_all("From", [])),
            "to": getaddresses(m.get_all("To", []) + m.get_all("Cc", [])),
            "date": dt,
            "is_sent": "Sent" in labels,
            "is_spam": "Spam" in labels,
            "self_hints": [a for _, a in getaddresses(m.get_all("Delivered-To", [])) if a],
            "body": _message_text(m, body_cap),
        }

    with open(path, "r", encoding="latin-1", errors="replace") as fh:
        buf = []
        size = 0
        for line in fh:
            if SEP.match(line):
                if buf:
                    yield normalize(buf)
                buf = []
                size = 0
                continue
            if body_cap is None or size < body_cap:   # headers + (capped) body
                buf.append(line)
                size += len(line)
        if buf:
            yield normalize(buf)


def _pst_record_value(message, entry_type):
    """Return the string value of a MAPI property on a pypff message, or ''."""
    try:
        nsets = message.number_of_record_sets
    except Exception:
        return ""
    for i in range(nsets):
        try:
            rs = message.get_record_set(i)
            for j in range(rs.number_of_entries):
                entry = rs.get_entry(j)
                if entry.entry_type == entry_type:
                    return (entry.get_data_as_string() or "").strip()
        except Exception:
            continue
    return ""


def _pst_date(message):
    for attr in ("client_submit_time", "delivery_time"):
        try:
            dt = getattr(message, attr)
        except Exception:
            dt = None
        if dt is not None:
            return _normalize_dt(dt)
    return None


def _pst_body(message, cap=BODY_CAP):
    """Return a pypff message's plain-text body (capped; None = full), or ''."""
    try:
        body = message.plain_text_body
    except Exception:
        body = None
    if not body:
        return ""
    if isinstance(body, bytes):
        body = body.decode("utf-8", "replace")
    return _cap(body, cap)


def _pst_message_fields(message, folder_name, body_cap=BODY_CAP):
    """Normalize one pypff message (in `folder_name`) into a message dict.

    Prefers the original RFC822 transport headers (present for most mail) and
    parses them like an mbox message; otherwise falls back to MAPI properties
    for the sender. Recipients are only recovered from transport headers, so
    Outlook-composed mail without headers contributes its sender only.
    """
    fname = (folder_name or "").strip().lower()
    is_sent = fname in PST_SENT_FOLDERS
    is_spam = fname in PST_SPAM_FOLDERS
    body = _pst_body(message, body_cap)

    try:
        headers = message.transport_headers
    except Exception:
        headers = None

    if headers:
        m = HeaderParser().parsestr(headers)
        try:
            dt = _normalize_dt(parsedate_to_datetime(m.get("Date")))
        except Exception:
            dt = None
        return {
            "subject": dec(m.get("Subject") or ""),
            "from": getaddresses(m.get_all("From", [])),
            "to": getaddresses(m.get_all("To", []) + m.get_all("Cc", [])),
            "date": dt,
            "is_sent": is_sent,
            "is_spam": is_spam,
            "self_hints": [],
            "body": body,
        }

    # MAPI fallback (no transport headers).
    name = getattr(message, "sender_name", None) or \
        _pst_record_value(message, _MAPI_SENDER_NAME)
    email = (_pst_record_value(message, _MAPI_SENDER_SMTP)
             or _pst_record_value(message, _MAPI_SENDER_EMAIL))
    if "@" not in email:
        email = ""
    return {
        "subject": getattr(message, "subject", None) or "",
        "from": [(name or "", email)],
        "to": [],
        "date": _pst_date(message),
        "is_sent": is_sent,
        "is_spam": is_spam,
        "self_hints": [],
        "body": body,
    }


def _walk_pst_folder(folder, body_cap=BODY_CAP):
    """Recursively yield normalized messages from a pypff folder."""
    try:
        name = folder.name or ""
    except Exception:
        name = ""
    try:
        n_msgs = folder.number_of_sub_messages
    except Exception:
        n_msgs = 0
    for i in range(n_msgs):
        try:
            message = folder.get_sub_message(i)
        except Exception:
            continue
        yield _pst_message_fields(message, name, body_cap)
    try:
        n_sub = folder.number_of_sub_folders
    except Exception:
        n_sub = 0
    for i in range(n_sub):
        try:
            sub = folder.get_sub_folder(i)
        except Exception:
            continue
        yield from _walk_pst_folder(sub, body_cap)


def iter_pst_messages(path, body_cap=BODY_CAP):
    """Yield normalized messages from an Outlook PST (requires pypff)."""
    try:
        import pypff
    except ImportError:
        sys.exit("error: reading .pst requires the 'libpff-python' package "
                 "(pip install libpff-python)")
    pst = pypff.file()
    pst.open(path)
    try:
        yield from _walk_pst_folder(pst.get_root_folder(), body_cap)
    finally:
        pst.close()


# ---- LLM corpus export ------------------------------------------------------
def _format_addrs(pairs):
    """Render (name, addr) pairs as a readable string, RFC2047 names decoded."""
    out = []
    for name, addr in pairs:
        name = dec(name).strip()
        if name and addr:
            out.append(f"{name} <{addr}>")
        elif addr:
            out.append(addr)
        elif name:
            out.append(name)
    return ", ".join(out)


def _is_noreply(pairs):
    """True if any sender address/name looks like a no-reply address."""
    for name, addr in pairs:
        blob = f"{name} {addr}".lower()
        if "noreply" in blob or "no-reply" in blob:
            return True
    return False


# Canonical extension appended to a bare -o name lacking one, per format.
_FMT_EXT = {"json": ".json", "csv": ".csv", "xlsx": ".xlsx", "outlook": ".csv",
            "vcard": ".vcf", "jsonl": ".jsonl"}


def _resolve_out(path, fmt):
    """Resolve an output path. A directory is rejected; a bare name without an
    extension gets the format's canonical suffix appended."""
    if os.path.isdir(path) or path.endswith(os.sep):
        sys.exit("error: -o must be a file path, not a directory: %s" % path)
    out = os.path.abspath(path)
    if not os.path.splitext(out)[1]:
        out += _FMT_EXT.get(fmt, "")
    return out


def write_contacts_as(path, rows, fmt):
    """Write contact rows in the given output format. The native full-column
    layout is json/csv/xlsx; outlook uses the Outlook column layout (.csv or
    .xlsx by extension); vcard is a vCard file."""
    if fmt == "vcard":
        write_vcards(path, rows)
    elif fmt == "outlook":
        (write_outlook_xlsx if path.lower().endswith(".xlsx")
         else write_outlook_csv)(path, rows)
    elif fmt == "xlsx":
        write_xlsx_rows(path, rows)
    elif fmt == "csv":
        write_csv_rows(path, rows)
    else:  # json
        write_json_rows(path, rows)


def _record_key(r):
    """Stable identity for a contact used to key a merge: its email, else its
    LinkedIn URL. This lets email-less LinkedIn-only contacts survive a merge
    (otherwise an email-keyed merge would silently drop them)."""
    email = (r.get("primary_email") or "").strip().lower()
    if email:
        return email
    url = (r.get("linkedin") or "").strip().rstrip("/").lower()
    return ("linkedin:" + url) if url else ""


def _read_existing_contacts(path, fmt):
    """Read an existing output file to fold an import into, as {record_key: row};
    a missing file yields {}."""
    if not os.path.isfile(path):
        return {}
    if fmt == "vcard":
        rows = parse_vcards(path)
    elif fmt == "outlook":
        rows = (parse_outlook_xlsx if path.lower().endswith(".xlsx")
                else parse_outlook_csv)(path)
    else:  # json / native csv / native xlsx
        rows = load_rows(path)
    out = {}
    for r in rows:
        key = _record_key(r)
        if key:
            out[key] = r
    return out


def _contact_sort_key(r):
    """Display order: named companies first, then by volume, then by name."""
    return (r["company"] == "", r["company"].lower(), -r["num_emails"],
            r["last_name"].lower(), r["first_name"].lower())


def _fold_into(existing, new_rows, force=False):
    """Fold new_rows into the existing {record_key: row} dict in place; returns
    (n_new, n_updated). With force=True, overlapping fields are overwritten."""
    n_new = n_updated = 0
    for row in new_rows:
        key = _record_key(row)
        if not key:
            continue   # no email and no LinkedIn URL: nothing to key it by
        if key in existing:
            merge_row(existing[key], row, force=force)
            n_updated += 1
        else:
            existing[key] = row
            n_new += 1
    return n_new, n_updated


def _run_date(args):
    """The date to stamp on imported records: an explicit --import-date override
    (for deterministic tests) or today's date, as an ISO string."""
    return getattr(args, "import_date", None) or date.today().isoformat()


def _stamp_import_date(rows, run_date):
    """Stamp import_date (and default linkedin) on freshly imported rows."""
    for r in rows:
        r.setdefault("linkedin", "")
        r["import_date"] = run_date
    return rows


def _name_key(first, last):
    """Normalized 'first last' key for matching (lowercased, punctuation/suffix
    stripped, e.g. 'Patil, PhD' -> 'patil'). Empty if no usable name."""
    f = re.sub(r"[^a-z0-9]+", " ", clean(first or "").lower()).strip()
    last = re.sub(r"[^a-z0-9]+", " ", clean(last or "").lower()).strip()
    # last name may carry a suffix token (phd/jr/...); keep the first token
    last = last.split(" ")[0] if last else ""
    f = f.split(" ")[0] if f else ""
    key = (f + " " + last).strip()
    return key if f and last else ""


def _new_linkedin_row(entry, run_date):
    """Build a fresh contact row from a LinkedIn entry (may be email-less)."""
    email = (entry.get("email") or "").strip().lower()
    return {
        "type": "", "friend": "",
        "last_name": clean(entry.get("last", "")),
        "first_name": clean(entry.get("first", "")),
        "title": entry.get("position", ""),
        "company": entry.get("company", ""),
        "phone": "", "address": "",
        "primary_email": email,
        "emails": [email] if email else [],
        "num_emails": 0, "num_sent": 0, "num_received": 0,
        "first_interaction": None, "last_interaction": None,
        "source": "linkedin",
        "linkedin": entry.get("url", ""),
        "import_date": run_date,
    }


def _fold_linkedin(rows, entries, run_date, source="linkedin"):
    """Fold LinkedIn entries into the contact rows in place. LinkedIn is the
    authority on current employer/title, so company/title are OVERWRITTEN on a
    match (unlike the generic merge). Matching priority: profile URL, then email,
    then normalized name. Ambiguous names (>1 match) are skipped. Unmatched
    entries are added as new contacts (keyed by URL when email-less).

    Returns (rows, n_enriched, n_added, n_ambiguous, n_skipped). Entries with
    neither an email nor a profile URL that match no existing contact are skipped
    (n_skipped): they have no key, so they would be lost on the next load and
    re-added on every run.
    """
    def norm_url(u):
        return (u or "").strip().rstrip("/").lower()

    by_email, by_url, by_name = {}, {}, defaultdict(list)

    def index(row):
        for e in [row.get("primary_email", "")] + list(row.get("emails") or []):
            e = (e or "").strip().lower()
            if e:
                by_email.setdefault(e, row)
        u = norm_url(row.get("linkedin"))
        if u:
            by_url.setdefault(u, row)
        k = _name_key(row.get("first_name"), row.get("last_name"))
        if k:
            by_name[k].append(row)

    for row in rows:
        index(row)

    n_enriched = n_added = n_ambiguous = n_skipped = 0
    for e in entries:
        url, email = norm_url(e.get("url")), (e.get("email") or "").strip().lower()
        match = None
        if url and url in by_url:
            match = by_url[url]
        elif email and email in by_email:
            match = by_email[email]
        else:
            cands = by_name.get(_name_key(e.get("first"), e.get("last")), [])
            if len(cands) == 1:
                match = cands[0]
            elif len(cands) > 1:
                n_ambiguous += 1
                continue
        if match is not None:
            if e.get("company"):
                match["company"] = e["company"]
            if e.get("position"):
                match["title"] = e["position"]
            if e.get("url"):
                match["linkedin"] = e["url"]
            match["import_date"] = run_date
            if email and email not in (a.lower() for a in match.get("emails", [])):
                match.setdefault("emails", []).append(email)
                if not match.get("primary_email"):
                    match["primary_email"] = email
            n_enriched += 1
        elif not email and not url:
            n_skipped += 1      # no email and no URL: nothing to key it by
        else:
            row = _new_linkedin_row(e, run_date)
            row["source"] = source
            rows.append(row)
            index(row)          # so duplicate LI rows don't double-add
            n_added += 1
    return rows, n_enriched, n_added, n_ambiguous, n_skipped


def dump_llm(src, out_path):
    """Stream a per-email JSONL corpus from a normalized-message source.

    One JSON object per line: subject/from/to/date/body. Skips spam, no-reply /
    automated-bulk senders, and empty bodies; strips quoted reply history from
    each body and de-duplicates identical bodies. Returns the records written.
    """
    outdir = os.path.dirname(out_path)
    if outdir:
        os.makedirs(outdir, exist_ok=True)
    n = kept = spam = bulk = empty = dup = 0
    seen = set()
    with open(out_path, "w", encoding="utf-8") as fh:
        for msg in src:
            n += 1
            if msg["is_spam"]:
                spam += 1
                continue
            from_emails = [a for _, a in msg["from"] if a]
            if _is_noreply(msg["from"]) or any(is_bot(a) for a in from_emails):
                bulk += 1
                continue
            body = _fresh_body(msg.get("body", ""))
            if not body:
                empty += 1
                continue
            digest = hashlib.md5(body.encode("utf-8", "replace")).digest()
            if digest in seen:
                dup += 1
                continue
            seen.add(digest)
            rec = {
                "subject": msg.get("subject", ""),
                "from": _format_addrs(msg["from"]),
                "to": _format_addrs(msg["to"]),
                "date": msg["date"].isoformat() if msg["date"] else "",
                "body": body,
            }
            fh.write(json.dumps(rec, ensure_ascii=False) + "\n")
            kept += 1
    sys.stderr.write(
        f"  dropped: {spam:,} spam, {bulk:,} bulk, {empty:,} empty, "
        f"{dup:,} duplicate (of {n:,} messages)\n")
    return kept


# ---- dedup ------------------------------------------------------------------
def _join_distinct(values, sep=SOURCE_SEP):
    """Join distinct non-empty values (each may itself already be sep-joined)."""
    out = []
    for v in values:
        for piece in (v or "").split(sep):
            piece = piece.strip()
            if piece and piece not in out:
                out.append(piece)
    return sep.join(out)


def _merge_group(rows):
    """Merge a list of same-name contact rows into one (counts sum; emails and
    sources union; dates widen; type/company/phone keep all distinct values;
    name casing and primary email come from the highest-volume row)."""
    winner = max(rows, key=lambda r: r["num_emails"])
    emails = []
    for r in rows:
        for e in r["emails"]:
            if e not in emails:
                emails.append(e)
    source = ""
    first = last = None
    for r in rows:
        source = _union_sources(source, r.get("source", ""))
        first = _merge_date(first, r.get("first_interaction"), newest=False)
        last = _merge_date(last, r.get("last_interaction"), newest=True)
    return {
        "type": _join_distinct(r["type"] for r in rows),
        "friend": _join_distinct(r["friend"] for r in rows),
        "last_name": winner["last_name"],
        "first_name": winner["first_name"],
        "title": _join_distinct(r["title"] for r in rows),
        "company": _join_distinct(r["company"] for r in rows),
        "phone": _join_distinct(r["phone"] for r in rows),
        "address": _join_distinct(r["address"] for r in rows),
        "primary_email": winner["primary_email"],
        "emails": emails,
        "num_emails": sum(r["num_emails"] for r in rows),
        "num_sent": sum(r["num_sent"] for r in rows),
        "num_received": sum(r["num_received"] for r in rows),
        "linkedin": next((r.get("linkedin") for r in rows if r.get("linkedin")), ""),
        "import_date": max((r.get("import_date") or "") for r in rows),
        "first_interaction": first,
        "last_interaction": last,
        "source": source,
    }


def dedup_contacts(rows):
    """Merge rows sharing a case-insensitive (first, last) name into one.

    Rows missing a first or last name cannot be keyed and pass through untouched.
    Output is sorted the same way as the importer.
    """
    groups = {}
    order = []
    passthrough = []
    for r in rows:
        first = r["first_name"].strip().lower()
        last = r["last_name"].strip().lower()
        if not first or not last:
            passthrough.append(r)
            continue
        key = (first, last)
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(r)
    merged = []
    for key in order:
        grp = groups[key]
        merged.append(grp[0] if len(grp) == 1 else _merge_group(grp))
    merged.extend(passthrough)
    merged.sort(key=lambda r: (
        r["company"] == "",
        r["company"].lower(),
        -r["num_emails"],
        r["last_name"].lower(),
        r["first_name"].lower(),
    ))
    return merged


# ---- reconcile (clean + cross-source merge) ---------------------------------
# Shared/generic mailbox local-parts: different people may all list these, so
# they are never used as a cross-record identity key, and are dropped as junk.
ROLE_LOCALPARTS = {
    "no-reply", "noreply", "no_reply", "donotreply", "do-not-reply",
    "info", "sales", "support", "admin", "administrator", "postmaster",
    "hello", "contact", "help", "office", "team", "marketing", "billing",
    "accounts", "accounting", "hr", "jobs", "careers", "press", "media",
    "webmaster", "abuse", "security", "privacy", "legal", "feedback",
    "newsletter", "news", "notifications", "notification", "service",
    "services", "enquiries", "inquiries", "mail", "email", "noreply-",
}


def _valid_email(addr):
    """Loose validity check: exactly one @, non-empty local, dotted domain."""
    addr = (addr or "").strip().lower()
    if addr.count("@") != 1 or " " in addr or "," in addr:
        return False
    local, dom = addr.split("@")
    return bool(local) and "." in dom and not dom.startswith(".") \
        and not dom.endswith(".")


def _is_role_address(addr):
    """True for shared/generic mailboxes (info@, sales@, no-reply@, ...)."""
    base = (addr or "").split("@", 1)[0].lower().strip().split("+", 1)[0]
    return base in ROLE_LOCALPARTS


def _is_free(addr):
    return (addr or "").split("@")[-1].lower() in FREE_PROVIDERS


def _mergeable_address(addr):
    """A valid, personal address usable as a cross-record identity key (not a
    free-provider, role/generic, or bot mailbox that different people may share)."""
    return (_valid_email(addr) and not _is_free(addr)
            and not _is_role_address(addr) and not is_bot(addr))


def _norm_company(name):
    return re.sub(r"[^a-z0-9]", "", (name or "").lower())


def _pick_primary(record):
    """Pick the primary email: prefer one whose domain maps to the record's
    company, else keep the current primary if still present, else emails[0]."""
    emails = record.get("emails") or []
    if not emails:
        return record.get("primary_email", "") or ""
    company = _norm_company(record.get("company", ""))
    if company:
        for e in emails:
            cf = _norm_company(company_from(e))
            if cf and (cf == company
                       or (len(cf) >= 4 and (cf in company or company in cf))):
                return e
    cur = record.get("primary_email") or ""
    if cur.lower() in (e.lower() for e in emails):
        return cur
    return emails[0]


def _normalize_name(name):
    """Trim/collapse whitespace; Title-case tokens that are all-upper or
    all-lower, leaving mixed-case names (McX, O'Brien, van) untouched."""
    out = []
    for tok in re.sub(r"\s+", " ", (name or "").strip()).split(" "):
        out.append(tok.capitalize() if tok and (tok.isupper() or tok.islower())
                   else tok)
    return " ".join(out).strip()


def _normalize_phone(phone, region=DEFAULT_REGION):
    """Normalize a stored phone to +E.164 if it parses to a valid number;
    otherwise leave it unchanged (do not discard data we cannot parse)."""
    phone = (phone or "").strip()
    if not phone:
        return ""
    try:
        num = phonenumbers.parse(phone, region)
        if phonenumbers.is_valid_number(num):
            return phonenumbers.format_number(
                num, phonenumbers.PhoneNumberFormat.E164)
    except Exception:
        pass
    return phone


def _reconcile_merge(group):
    """Merge a group of duplicate records into one. company/title come from a
    LinkedIn-sourced member if any; other fields from the newest-interaction
    member (fallback import_date); emails/sources union, counts sum, dates widen."""
    if len(group) == 1:
        return dict(group[0])
    base = max(group, key=lambda r: (r.get("last_interaction") or "",
                                     r.get("import_date") or ""))
    merged = dict(base)

    def first_nonblank(field):
        return next((r[field] for r in group if str(r.get(field) or "").strip()),
                    merged.get(field, ""))

    li = next((r for r in group if str(r.get("linkedin") or "").strip()), None)
    for f in ("company", "title"):
        if li and str(li.get(f) or "").strip():
            merged[f] = li[f]
        elif not str(merged.get(f) or "").strip():
            merged[f] = first_nonblank(f)
    for f in ("first_name", "last_name", "phone", "address", "type", "friend"):
        if not str(merged.get(f) or "").strip():
            merged[f] = first_nonblank(f)

    emails = []
    for r in [base] + list(group):
        for e in [r.get("primary_email", "")] + list(r.get("emails") or []):
            e = (e or "").strip()
            if e and e.lower() not in (x.lower() for x in emails):
                emails.append(e)
    merged["emails"] = emails
    merged["num_sent"] = sum(int(r.get("num_sent") or 0) for r in group)
    merged["num_received"] = sum(int(r.get("num_received") or 0) for r in group)
    merged["num_emails"] = merged["num_sent"] + merged["num_received"]
    fi = [r.get("first_interaction") for r in group if r.get("first_interaction")]
    la = [r.get("last_interaction") for r in group if r.get("last_interaction")]
    merged["first_interaction"] = min(fi) if fi else None
    merged["last_interaction"] = max(la) if la else None
    src = ""
    for r in group:
        src = _union_sources(src, r.get("source", "") or "")
    merged["source"] = src
    merged["linkedin"] = next(
        (r["linkedin"] for r in group if str(r.get("linkedin") or "").strip()), "")
    merged["import_date"] = max((r.get("import_date") or "") for r in group)
    return merged


def _merge_by_shared_email(rows):
    """Union-find merge of records that share a mergeable (personal) email."""
    parent = list(range(len(rows)))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    seen = {}
    for i, r in enumerate(rows):
        for e in r.get("emails") or []:
            if _mergeable_address(e):
                el = e.lower()
                if el in seen:
                    parent[find(i)] = find(seen[el])
                else:
                    seen[el] = i
    groups = defaultdict(list)
    for i in range(len(rows)):
        groups[find(i)].append(rows[i])
    return [_reconcile_merge(g) for g in groups.values()]


def _merge_by_name(rows):
    """Merge records sharing a lowercased (first, last) name (the --dedup pass,
    using reconcile's single-winner merge)."""
    groups, order, passthrough = {}, [], []
    for r in rows:
        first = (r.get("first_name") or "").strip().lower()
        last = (r.get("last_name") or "").strip().lower()
        if not first or not last:
            passthrough.append(r)
            continue
        key = (first, last)
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(r)
    out = [_reconcile_merge(groups[k]) for k in order]
    out.extend(passthrough)
    return out


def reconcile_contacts(rows):
    """Clean and merge a contacts DB: drop junk addresses, merge duplicates by
    shared email and by name, recompute derived fields, pick the best primary
    email, and normalize names/phones. Records left with no email and no LinkedIn
    URL are dropped. Returns sorted rows."""
    rows = [dict(r) for r in rows]
    # 1. drop junk addresses; re-point primary if it was dropped
    for r in rows:
        deduped, seen = [], set()
        for e in r.get("emails") or []:
            el = (e or "").strip().lower()
            if (el and el not in seen and _valid_email(el)
                    and not is_bot(el) and not _is_role_address(el)):
                seen.add(el)
                deduped.append(el)
        r["emails"] = deduped
        pe = (r.get("primary_email") or "").lower()
        r["primary_email"] = pe if pe in seen else (deduped[0] if deduped else "")
    # 2. merge by shared personal email, then 3. by name
    rows = _merge_by_shared_email(rows)
    rows = _merge_by_name(rows)
    # 4-7. per-record cleanup
    out = []
    for r in rows:
        r["num_emails"] = int(r.get("num_sent") or 0) + int(r.get("num_received") or 0)
        if not str(r.get("company") or "").strip() and r.get("primary_email"):
            r["company"] = company_from(r["primary_email"])
        r["primary_email"] = _pick_primary(r)
        r["first_name"] = _normalize_name(r.get("first_name", ""))
        r["last_name"] = _normalize_name(r.get("last_name", ""))
        r["phone"] = _normalize_phone(r.get("phone", ""))
        if r.get("primary_email") or str(r.get("linkedin") or "").strip():
            out.append(r)
    out.sort(key=_contact_sort_key)
    return out


# ---- vCard export -----------------------------------------------------------
def _vcard_escape(value):
    """Escape a value for a vCard 3.0 text field (RFC 2426)."""
    return (str(value).replace("\\", "\\\\").replace("\n", "\\n")
            .replace(",", "\\,").replace(";", "\\;"))


def _vcard_fold(line):
    """Fold a vCard line to <=75 octets, continuations led by a single space.

    Splits on character boundaries (never inside a multi-byte UTF-8 char) and
    joins physical lines with CRLF, as Google Contacts expects.
    """
    physical = []
    cur, cur_bytes = "", 0
    for ch in line:
        b = len(ch.encode("utf-8"))
        if cur_bytes + b > 75:
            physical.append(cur)
            cur, cur_bytes = " ", 1     # continuation leading space
        cur += ch
        cur_bytes += b
    physical.append(cur)
    return "\r\n".join(physical)


def _contact_vcard(row):
    """Return the folded vCard 3.0 lines for one contact row (Gmail style)."""
    first = (row.get("first_name") or "").strip()
    last = (row.get("last_name") or "").strip()
    fn = (first + " " + last).strip() or (row.get("primary_email") or "")
    props = [
        "BEGIN:VCARD",
        "VERSION:3.0",
        "FN:" + _vcard_escape(fn),
        "N:%s;%s;;;" % (_vcard_escape(last), _vcard_escape(first)),
    ]
    if row.get("company"):
        props.append("ORG:" + _vcard_escape(row["company"]))
    if row.get("title"):
        props.append("TITLE:" + _vcard_escape(row["title"]))

    primary = row.get("primary_email") or ""
    ordered = ([primary] if primary else [])
    ordered += [e for e in (row.get("emails") or []) if e and e != primary]
    for i, email in enumerate(ordered):
        typ = "INTERNET,PREF" if i == 0 else "INTERNET"
        props.append("EMAIL;TYPE=%s:%s" % (typ, email))

    if row.get("phone"):
        props.append("TEL;TYPE=VOICE:" + row["phone"])
    if row.get("address"):
        addr = _vcard_escape(row["address"])
        props.append("ADR;TYPE=WORK:;;%s;;;;" % addr)
        props.append("LABEL;TYPE=WORK:" + addr)

    note = "emails %s (sent %s, received %s)" % (
        row.get("num_emails", 0), row.get("num_sent", 0),
        row.get("num_received", 0))
    if row.get("first_interaction") or row.get("last_interaction"):
        note += "; %s..%s" % (row.get("first_interaction") or "?",
                              row.get("last_interaction") or "?")
    if row.get("source"):
        note += "; source: " + row["source"]
    props.append("NOTE:" + _vcard_escape(note))

    cats = [c for c in (row.get("type") or "",
                        "friend" if row.get("friend") else "") if c]
    if cats:
        props.append("CATEGORIES:" + ",".join(_vcard_escape(c) for c in cats))
    props.append("END:VCARD")
    return [_vcard_fold(p) for p in props]


def _vcard_text(rows):
    """Render one or more contact rows as CRLF-delimited, folded vCard 3.0."""
    lines = []
    for r in rows:
        lines.extend(_contact_vcard(r))
    return "\r\n".join(lines) + "\r\n"


def write_vcards(path, rows):
    """Write all contacts into a single multi-card vCard file (Gmail-style)."""
    _write_atomic(path, lambda out: out.write(_vcard_text(rows)))


def _vcard_unescape(value):
    """Reverse vCard text escaping (\\\\ \\n \\, \\;)."""
    out, i = [], 0
    while i < len(value):
        ch = value[i]
        if ch == "\\" and i + 1 < len(value):
            nxt = value[i + 1]
            out.append("\n" if nxt in ("n", "N") else nxt)
            i += 2
        else:
            out.append(ch)
            i += 1
    return "".join(out)


def _vcard_split(value, sep):
    """Split on unescaped `sep`; returns still-escaped pieces."""
    parts, cur, i = [], [], 0
    while i < len(value):
        ch = value[i]
        if ch == "\\" and i + 1 < len(value):
            cur.append(value[i:i + 2])
            i += 2
        elif ch == sep:
            parts.append("".join(cur))
            cur = []
            i += 1
        else:
            cur.append(ch)
            i += 1
    parts.append("".join(cur))
    return parts


def _unfold_vcard_lines(text):
    """Unfold RFC 2426 continuation lines (leading space/tab)."""
    lines = []
    for ln in text.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        if ln[:1] in (" ", "\t") and lines:
            lines[-1] += ln[1:]
        else:
            lines.append(ln)
    return lines


def _vcard_to_row(card, source):
    """Build a contact row dict from accumulated vCard properties."""
    n = card["n"]
    if n and len(n) >= 2 and (n[0].strip() or n[1].strip()):
        last, first = n[0].strip(), n[1].strip()
    else:
        first, last = split_name(card["fn"], "")
    primary = card["pref_email"] or (card["emails"][0] if card["emails"] else "")
    emails = ([primary] if primary else [])
    emails += [e for e in card["emails"] if e and e != primary and e not in emails]
    ctype, friend = "", ""
    for c in card["categories"]:
        if c.lower() in TYPE_VALUES and not ctype:
            ctype = c.lower()
        if c.lower() == "friend":
            friend = "Y"
    if not (first or last or primary):
        return None
    return _normalize_row({
        "type": ctype, "friend": friend, "first_name": first, "last_name": last,
        "title": card["title"], "company": card["org"], "phone": card["tel"],
        "address": card["label"] or card["adr"], "primary_email": primary,
        "emails": emails, "source": source,
    })


def parse_vcards(path):
    """Parse a vCard (.vcf) file into a list of contact row dicts."""
    with open(path, "r", encoding="utf-8", errors="replace") as fh:
        text = fh.read()
    source = os.path.basename(path)
    rows, card = [], None
    for line in _unfold_vcard_lines(text):
        s = line.strip()
        if s.upper() == "BEGIN:VCARD":
            card = {"emails": [], "pref_email": "", "tel": "", "n": None,
                    "fn": "", "org": "", "title": "", "adr": "", "label": "",
                    "categories": []}
            continue
        if s.upper() == "END:VCARD":
            if card is not None:
                row = _vcard_to_row(card, source)
                if row:
                    rows.append(row)
            card = None
            continue
        if card is None or ":" not in line:
            continue
        left, _, value = line.partition(":")
        bits = left.split(";")
        name, params = bits[0].upper(), ";".join(bits[1:]).upper()
        if name == "FN":
            card["fn"] = _vcard_unescape(value)
        elif name == "N":
            card["n"] = [_vcard_unescape(c) for c in _vcard_split(value, ";")]
        elif name == "ORG":
            comps = _vcard_split(value, ";")
            card["org"] = _vcard_unescape(comps[0]) if comps else ""
        elif name == "TITLE":
            card["title"] = _vcard_unescape(value)
        elif name == "EMAIL":
            addr = _vcard_unescape(value).strip()
            if addr:
                card["emails"].append(addr)
                if "PREF" in params:
                    card["pref_email"] = addr
        elif name == "TEL":
            if not card["tel"]:
                card["tel"] = _vcard_unescape(value).strip()
        elif name == "ADR":
            comps = [_vcard_unescape(c).strip() for c in _vcard_split(value, ";")]
            card["adr"] = ", ".join(c for c in comps if c)
        elif name == "LABEL":
            card["label"] = _vcard_unescape(value).replace("\n", ", ").strip(", ")
        elif name == "CATEGORIES":
            for c in _vcard_split(value, ","):
                c = _vcard_unescape(c).strip()
                if c:
                    card["categories"].append(c)
    return rows


# ---- commands ---------------------------------------------------------------
def _merge_and_write(args, new_rows, ofmt):
    """Fold new_rows into the existing output DB and write it back. An import
    never wipes the target: existing contacts are kept, counts are overwritten
    with this batch, emails and sources union, the date range widens. Hand-edited
    text fields are preserved unless --force overwrites overlapping fields. A
    missing output file is created fresh."""
    cpath = _resolve_out(args.output, ofmt)
    outdir = os.path.dirname(cpath)
    if outdir:
        os.makedirs(outdir, exist_ok=True)
    merged = _read_existing_contacts(cpath, ofmt)
    n_existing = len(merged)
    n_new, n_updated = _fold_into(merged, new_rows, force=args.force)
    rows = sorted(merged.values(), key=_contact_sort_key)
    write_contacts_as(cpath, rows, ofmt)
    sys.stderr.write(
        f"\nWrote {cpath}\n"
        f"  {n_existing:,} existing + {n_new:,} new "
        f"({n_updated:,} updated) = {len(rows):,} contacts.\n")


def cmd_import_linkedin(args, ofmt, run_date):
    """Import a LinkedIn Connections CSV, folding into the existing output DB
    (overwriting company/title -- LinkedIn is the authority -- adding the profile
    URL and any new connections). A missing output file is created fresh."""
    entries = parse_linkedin_csv(args.input)
    source = os.path.basename(args.input)
    cpath = _resolve_out(args.output, ofmt)
    outdir = os.path.dirname(cpath)
    if outdir:
        os.makedirs(outdir, exist_ok=True)
    existing = load_rows(cpath)
    n_existing = len(existing)
    rows, n_enriched, n_added, n_ambiguous, n_skipped = _fold_linkedin(
        existing, entries, run_date, source=source)
    rows = sorted(rows, key=_contact_sort_key)
    write_contacts_as(cpath, rows, ofmt)
    sys.stderr.write(
        f"\nWrote {cpath} from {source}\n"
        f"  {len(entries):,} LinkedIn rows: {n_enriched:,} enriched, "
        f"{n_added:,} added, {n_ambiguous:,} ambiguous-skipped, "
        f"{n_skipped:,} skipped (no email/URL).\n"
        f"  {n_existing:,} existing -> {len(rows):,} contacts.\n")


def cmd_dump_llm(args, ifmt):
    """Dump a per-email JSONL corpus from an mbox/PST; no DB. Bodies are capped
    at --max-body bytes (0 = unlimited) so attachment blobs don't dominate."""
    path = args.input
    if not os.path.isfile(path):
        sys.exit("error: input not found: %s" % path)
    cap = args.max_body or None   # 0 -> unlimited
    src = (iter_pst_messages(path, body_cap=cap) if ifmt == "pst"
           else iter_mbox_messages(path, body_cap=cap))
    out_path = _resolve_out(args.output, "jsonl")
    n = dump_llm(src, out_path)
    sys.stderr.write(f"Wrote {n:,} email records to {out_path}\n")


def cmd_import(args, ifmt, ofmt):
    """Import a mailbox, vCard, or Outlook CSV (args.input) and write the
    contacts in ofmt (a .json DB, .csv, Outlook CSV, or vCard)."""
    path = args.input
    # Self addresses are auto-detected: the mbox Delivered-To header / the From
    # of Sent-folder (or Sent-labeled) mail.
    self_set = set()
    run_date = _run_date(args)   # stamped as import_date on imported records

    if not os.path.isfile(path):
        sys.exit("error: input not found: %s" % path)

    # LinkedIn Connections export: fold into the existing output DB (overwrite
    # company/title, add the profile URL and any new connections).
    if ifmt == "linkedin":
        return cmd_import_linkedin(args, ofmt, run_date)

    # Outlook CSV/XLSX: read the Outlook column layout straight into rows.
    if ifmt == "outlook":
        new_rows = (parse_outlook_xlsx if path.lower().endswith(".xlsx")
                    else parse_outlook_csv)(path)
        _stamp_import_date(new_rows, run_date)
        sys.stderr.write(
            f"Parsed {len(new_rows):,} contacts from {os.path.basename(path)}.\n")
        _merge_and_write(args, new_rows, ofmt)
        return

    # vCard input: contacts come straight from the cards (no message pipeline).
    if ifmt == "vcard":
        new_rows = parse_vcards(path)
        _stamp_import_date(new_rows, run_date)
        sys.stderr.write(
            f"Parsed {len(new_rows):,} contacts from {os.path.basename(path)}.\n")
        _merge_and_write(args, new_rows, ofmt)
        return

    pst = ifmt == "pst"
    blacklist = set()
    if args.blacklist:
        blacklist = load_domain_files(args.blacklist, "blacklist")

    recs = defaultdict(Rec)  # primary email -> Rec
    n_msgs = 0
    n_skip_spam = 0

    # Pick the reader by extension: .pst -> Outlook, otherwise mbox.
    src = iter_pst_messages(path) if pst else iter_mbox_messages(path)
    for msg in src:
        n_msgs += 1
        if not _ingest_message(msg, recs, self_set,
                               include_cc=not getattr(args, "no_cc", False)):
            n_skip_spam += 1
        if n_msgs % 50000 == 0:
            sys.stderr.write(f"  parsed {n_msgs:,} messages\n")
            sys.stderr.flush()

    sys.stderr.write(f"Done parsing: {n_msgs:,} messages, {n_skip_spam:,} spam skipped.\n")
    sys.stderr.write(f"Self addresses: {sorted(self_set)}\n")

    # ---- drop self + blacklist + bots ---------------------------------------
    for a in list(recs):
        if a in self_set:
            del recs[a]
    blacklisted = {a for a in recs if is_blacklisted(a, blacklist)}
    for a in blacklisted:
        del recs[a]
    if blacklist:
        sys.stderr.write(
            f"Filtered {len(blacklisted):,} blacklisted addresses "
            f"({len(blacklist):,} domains).\n")
    bots = {a for a in recs if is_bot(a)}
    for a in bots:
        del recs[a]
    sys.stderr.write(f"Filtered {len(bots):,} automated/bulk addresses.\n")

    # ---- merge by display name ---------------------------------------------
    def canon_name(r):
        if not r.names:
            return None
        nm = max(r.names, key=r.names.get)
        return re.sub(r'\s+', ' ', nm.lower().replace(",", " ")).strip()

    by_name = defaultdict(list)
    no_name = []
    for addr, r in recs.items():
        key = canon_name(r)
        if key:
            by_name[key].append(addr)
        else:
            no_name.append(addr)

    people = []

    def build(addrs):
        merged = Rec()
        for a in addrs:
            r = recs[a]
            for e, c in r.emails.items():
                merged.emails[e] += c
            for nm, c in r.names.items():
                merged.names[nm] += c
            for ph, c in r.phones.items():
                merged.phones[ph] += c
            merged.num_sent += r.num_sent
            merged.num_recv += r.num_recv
            merged.touch(r.first)
            merged.touch(r.last)
        # primary email = most-used address
        primary = max(merged.emails, key=merged.emails.get)
        display = max(merged.names, key=merged.names.get) if merged.names else ""
        first, last = split_name(display, primary)
        # phone = most-frequently-seen signature number, if any
        phone = max(merged.phones, key=merged.phones.get) if merged.phones else ""
        return {
            "first_name": first,
            "last_name": last,
            "company": company_from(primary),
            "phone": phone,
            "primary_email": primary,
            "emails": sorted(merged.emails, key=lambda e: -merged.emails[e]),
            "display_name": display,
            "num_emails": merged.num_sent + merged.num_recv,
            "num_sent": merged.num_sent,
            "num_received": merged.num_recv,
            "first_interaction": merged.first.date().isoformat() if merged.first else None,
            "last_interaction": merged.last.date().isoformat() if merged.last else None,
        }

    for key, addrs in by_name.items():
        people.append(build(addrs))
    for addr in no_name:
        people.append(build([addr]))

    # ---- final contact filter: keep anyone with a full name. This includes
    # people you only shared a thread with (To/Cc co-recipients of mail you
    # received), who have no sent/received count.
    n_built = len(people)
    people = [p for p in people if p["first_name"] and p["last_name"]]
    n_kept = len(people)
    sys.stderr.write(
        f"Contacts: {n_built:,} built -> {n_kept:,} with full name.\n")

    source = os.path.basename(path)
    new_rows = [person_to_row(p, source) for p in people]
    _stamp_import_date(new_rows, run_date)
    _merge_and_write(args, new_rows, ofmt)


def cmd_export(args, ofmt):
    """Export matching records as CSV/XLSX, Outlook CSV/XLSX, or vCard."""
    if args.type:
        bad = [t for t in (_csv_set(args.type) or set()) if t not in TYPE_VALUES]
        if bad:
            sys.exit("error: invalid --type value(s) %s; legal values: %s"
                     % (", ".join(sorted(bad)), ", ".join(TYPE_VALUES)))
    whitelist, blacklist = load_domain_filters(args)
    contacts = load_rows(args.input)
    crit = build_criteria(args)
    selected = [c for c in contacts if matches(c, crit)]
    selected, n_wl, n_bl = select_by_domains(selected, whitelist, blacklist)
    if whitelist is not None or blacklist:
        sys.stderr.write(
            f"Domain filter: dropped {n_wl:,} not whitelisted, "
            f"{n_bl:,} blacklisted.\n")

    out = _resolve_out(args.output, ofmt)
    outdir = os.path.dirname(out)
    if outdir:
        os.makedirs(outdir, exist_ok=True)
    write_contacts_as(out, selected, ofmt)
    sys.stderr.write(
        f"Exported {len(selected):,}/{len(contacts):,} contacts to {out}\n")


def cmd_db(args, ofmt):
    """Fold a native contacts DB (json/csv/xlsx) into the existing output DB,
    never wiping it (--force overwrites overlapping fields), optionally filtering
    the input by domain and/or --dedup-ing the result. A missing output file is
    created fresh. When input and output are the same file, this normalizes it."""
    whitelist, blacklist = load_domain_filters(args)
    rows = load_rows(args.input)
    if not rows:
        sys.exit("error: no contacts found in %s" % args.input)
    rows, n_wl, n_bl = select_by_domains(rows, whitelist, blacklist)
    if whitelist is not None or blacklist:
        sys.stderr.write(
            f"Domain filter: dropped {n_wl:,} not whitelisted, "
            f"{n_bl:,} blacklisted.\n")
    out_path = _resolve_out(args.output, ofmt)
    outdir = os.path.dirname(out_path)
    if outdir:
        os.makedirs(outdir, exist_ok=True)

    # Fold the (filtered) input into the existing output DB; a missing output
    # starts empty (fresh write).
    existing = _read_existing_contacts(out_path, ofmt)
    n_existing = len(existing)
    n_new, n_updated = _fold_into(existing, rows, force=args.force)
    rows = list(existing.values())
    n_folded = len(rows)
    if args.reconcile:
        rows = reconcile_contacts(rows)
        extra = " then reconciled"
    elif args.dedup:
        rows = dedup_contacts(rows)
        extra = " then deduped"
    else:
        extra = ""
    rows = sorted(rows, key=_contact_sort_key)
    write_contacts_as(out_path, rows, ofmt)

    note = (f" ({n_folded - len(rows):,} merged/dropped)"
            if (args.reconcile or args.dedup) else "")
    sys.stderr.write(
        f"\nWrote {out_path}{extra}\n"
        f"  {n_existing:,} existing + {n_new:,} new from "
        f"{os.path.basename(args.input)} = {len(rows):,} contacts "
        f"({n_updated:,} updated){note}.\n")


# Recognized file formats. JSON is the native database; the rest are import
# sources / export targets. "outlook" is the Outlook/Google CSV column layout.
FORMATS = ["json", "csv", "xlsx", "outlook", "vcard", "linkedin", "mbox", "pst",
           "jsonl"]

# How a file's format is inferred from its extension (unless --iformat/--oformat
# overrides it). .csv/.xlsx default to the native layout; "outlook" must be asked
# for. json/csv/xlsx are the interchangeable native database formats.
_EXT_FORMAT = {".json": "json", ".csv": "csv", ".xlsx": "xlsx",
               ".vcf": "vcard", ".vcd": "vcard",
               ".mbox": "mbox", ".pst": "pst", ".jsonl": "jsonl"}
DB_FORMATS = ("json", "csv", "xlsx")


def resolve_format(path, override):
    """Resolve a file's format: an explicit --iformat/--oformat wins, otherwise
    infer from the extension. Returns a FORMATS value, or None if unknown."""
    if override:
        return override
    return _EXT_FORMAT.get(os.path.splitext(path)[1].lower())


def parse_args(argv=None):
    p = argparse.ArgumentParser(
        prog="mc",
        description="mailcompiler: build and query a contacts database. The "
                    "operation is inferred from the -i/-o formats: a mailbox, "
                    "vCard, Outlook CSV, or LinkedIn export imports into a "
                    "contacts DB; a JSON input exports (-o .csv/.vcf) or "
                    "deduplicates (-o .json --dedup). Imports and DB writes "
                    "always fold into the existing -o (never wiping it); pass "
                    "--force to overwrite overlapping fields.")
    p.add_argument("-i", "--input", dest="input", required=True,
                   help="input path: a mailbox (.mbox/.pst), a vCard "
                        "(.vcf/.vcd), an Outlook CSV (--iformat outlook), or a "
                        "contacts .json")
    p.add_argument("-o", "--output", dest="output", required=True,
                   help="output path: a .json contacts DB, a .csv/.vcf export, "
                        "or a .jsonl corpus (with --llm)")
    p.add_argument("--iformat", choices=FORMATS,
                   help="force the input format instead of inferring it from "
                        "the extension; 'outlook' reads an Outlook/Google CSV")
    p.add_argument("--oformat", choices=FORMATS,
                   help="force the output format instead of inferring it from "
                        "the extension; 'outlook' writes Outlook's CSV layout")
    p.add_argument("--dedup", action="store_true",
                   help="merge contacts sharing a first+last name (json -> json)")
    p.add_argument("--reconcile", action="store_true",
                   help="clean and merge records (json -> json): drop junk/role "
                        "addresses, merge duplicates by email and by name, "
                        "recompute fields, and pick the best primary email. "
                        "A superset of --dedup.")
    p.add_argument("--force", action="store_true",
                   help="when an imported record overlaps an existing one, "
                        "overwrite the existing text fields (company, title, "
                        "name, ...) with the incoming values; by default existing "
                        "(hand-edited) values are kept")
    p.add_argument("--llm", action="store_true",
                   help="dump a per-email JSONL corpus (subject/from/to/date/"
                        "body) from an mbox/PST instead of building the DB")
    p.add_argument("--max-body", dest="max_body", type=int, default=262144,
                   metavar="BYTES",
                   help="--llm: cap each message body to BYTES (default 262144; "
                        "0 = unlimited) so attachment blobs do not dominate")
    p.add_argument("--no-cc", dest="no_cc", action="store_true",
                   help="when importing a mailbox, do NOT bring in the other "
                        "To/Cc recipients of mail you received; keep only direct "
                        "senders and the recipients of your sent mail (less noise)")
    # Override the import_date stamp (default: today). Hidden; used by tests.
    p.add_argument("--import-date", dest="import_date", metavar="YYYY-MM-DD",
                   help=argparse.SUPPRESS)
    p.add_argument("--whitelist", dest="whitelist", metavar="PATH",
                   nargs="+", action="extend",
                   help="keep only contacts whose email domain matches an entry "
                        "in these files (one domain per line; '#' comments and "
                        "blank lines ignored; subdomains match too). Accepts "
                        "multiple files (unioned) and may be repeated.")
    p.add_argument("--blacklist", dest="blacklist", metavar="PATH",
                   nargs="+", action="extend",
                   help="drop contacts/addresses whose email domain matches an "
                        "entry in these files (one per line; '#' comments and "
                        "blank lines ignored; subdomains match too). Accepts "
                        "multiple files (unioned) and may be repeated.")
    # Export filters (apply when the operation resolves to an export).
    p.add_argument("--type", dest="type",
                   help="match contact type against any of LIST (%s)"
                        % "/".join(TYPE_VALUES))
    p.add_argument("--company", help="match company against any of LIST")
    p.add_argument("--first-name", dest="first_name",
                   help="match first name against any of LIST")
    p.add_argument("--last-name", dest="last_name",
                   help="match last name against any of LIST")
    p.add_argument("--email-domain", dest="email_domain",
                   help="match primary email domain against any of LIST")
    p.add_argument("--min-emails", type=int, help="minimum num_emails")
    p.add_argument("--max-emails", type=int, help="maximum num_emails")
    p.add_argument("--min-sent", type=int, help="minimum num_sent")
    p.add_argument("--max-sent", type=int, help="maximum num_sent")
    p.add_argument("--min-received", type=int, help="minimum num_received")
    p.add_argument("--max-received", type=int, help="maximum num_received")
    p.add_argument("--last-after", dest="last_after", metavar="YYYY-MM-DD",
                   help="last_interaction on or after this date")
    p.add_argument("--last-before", dest="last_before", metavar="YYYY-MM-DD",
                   help="last_interaction on or before this date")
    p.add_argument("--first-after", dest="first_after", metavar="YYYY-MM-DD",
                   help="first_interaction on or after this date")
    p.add_argument("--first-before", dest="first_before", metavar="YYYY-MM-DD",
                   help="first_interaction on or before this date")
    return p.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)
    ifmt = resolve_format(args.input, args.iformat)
    ofmt = resolve_format(args.output, args.oformat)
    if ifmt is None:
        sys.exit("error: cannot determine the format of input %s; "
                 "pass --iformat" % args.input)
    if ofmt is None:
        sys.exit("error: cannot determine the format of output %s; "
                 "pass --oformat" % args.output)

    # --llm: per-email JSONL dump from a mailbox (no contacts DB).
    if args.llm:
        if ifmt not in ("mbox", "pst"):
            sys.exit("error: --llm requires an mbox or PST input")
        if ofmt != "jsonl":
            sys.exit("error: --llm writes a .jsonl corpus, so -o must be .jsonl")
        if args.dedup or args.reconcile:
            sys.exit("error: --llm cannot be combined with --dedup/--reconcile")
        return cmd_dump_llm(args, ifmt)

    # Import: a mailbox / vCard / Outlook CSV / LinkedIn export builds contacts,
    # written in the output format (.json DB, .csv, Outlook CSV, or .vcf).
    if ifmt in ("mbox", "pst", "vcard", "outlook", "linkedin"):
        if ofmt == "jsonl":
            sys.exit("error: a .jsonl corpus is produced only with --llm")
        if ofmt not in DB_FORMATS and ifmt == "linkedin":
            sys.exit("error: a LinkedIn import writes a contacts database; "
                     "-o must be .json/.csv/.xlsx (not %s)" % ofmt)
        if args.dedup or args.reconcile:
            sys.exit("error: --dedup/--reconcile apply to a json -> json "
                     "database, not an import; import to .json first, then run it")
        return cmd_import(args, ifmt, ofmt)

    # Native contacts DB input (json/csv/xlsx -- interchangeable layouts).
    if ifmt in DB_FORMATS:
        if ofmt == "jsonl":
            sys.exit("error: a .jsonl corpus is produced only with --llm")
        if args.dedup or args.reconcile:
            if ofmt not in DB_FORMATS:
                sys.exit("error: --dedup/--reconcile produce a native database; "
                         "-o must be .json/.csv/.xlsx (not %s)" % ofmt)
            return cmd_db(args, ofmt)
        # DB -> .json folds into the existing DB (never wipes it); DB -> other
        # native formats / vCard / Outlook is a (filtered) export view.
        if ofmt == "json":
            return cmd_db(args, ofmt)
        return cmd_export(args, ofmt)   # csv/xlsx/outlook/vcard, with filters

    sys.exit("error: unsupported input format: %s" % ifmt)


if __name__ == "__main__":
    main()
